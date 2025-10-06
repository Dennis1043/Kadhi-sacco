from flask import Flask, send_from_directory, request, jsonify
from flask_cors import CORS
import mysql.connector
import traceback
from io import BytesIO
from flask import send_file
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from flask import Flask, request, jsonify, send_file
import mysql.connector
from werkzeug.security import generate_password_hash, check_password_hash
import firebase_admin
from firebase_admin import credentials, auth
import pandas as pd
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import csv

from flask import Flask, send_file
from fpdf import FPDF
import io, datetime
app = Flask(__name__, static_folder=".", static_url_path="")
CORS(app)

# ========================
# DATABASE CONFIG
# ========================
db_config = {
    "host": "localhost",
    "user": "root",
    "password": "@Dennis1043",
    "database": "kadhi"
}
def get_connection():
    return mysql.connector.connect(**db_config)

def get_mysql_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="@Dennis1043",
        database="kadhi"
    )

def get_db_connection():
    return mysql.connector.connect(**db_config)



def get_user_id_from_firebase(firebase_uid):
    """Return numeric user_id from firebase_uid"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id FROM users WHERE firebase_uid=%s", (firebase_uid,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()
    if not user:
        return None
    return user["id"]

def get_member_data(uid):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get user info (from users table)
    cursor.execute("SELECT * FROM users WHERE firebase_uid=%s", (uid,))
    user = cursor.fetchone()

    if not user:
        conn.close()
        return None

    user_id = user["id"]

    # Try to also check members table (if exists)
    cursor.execute("SELECT * FROM members WHERE firebase_uid=%s", (uid,))
    member = cursor.fetchone()

    # Use info from users primarily, fallback to members if needed
    name = user.get("name") or (member["name"] if member else "")
    email = user.get("email") or (member["email"] if member else "")
    phone = user.get("phone") or (member["phone"] if member else "")
    created_at = user.get("created_at") or (member["created_at"] if member else "")

    # Get contributions
    cursor.execute("SELECT * FROM contributions WHERE user_id=%s", (user_id,))
    contributions = cursor.fetchall() or []

    # Get loans
    cursor.execute("SELECT * FROM loans WHERE user_id=%s", (user_id,))
    loans = cursor.fetchall() or []

    # Get repayments
    cursor.execute("SELECT * FROM repayments WHERE user_id=%s", (user_id,))
    repayments = cursor.fetchall() or []

    # Get fines
    cursor.execute("SELECT * FROM fines WHERE user_id=%s", (user_id,))
    fines = cursor.fetchall() or []

    cursor.close()
    conn.close()

    return {
        "id": user_id,
        "name": name,
        "email": email,
        "phone": phone,
        "created_at": created_at,
        "contributions": contributions,
        "loans": loans,
        "repayments": repayments,
        "fines": fines
    }



from flask import send_file
from datetime import datetime
import io
import csv
import openpyxl
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib.styles import getSampleStyleSheet

@app.route("/download/<filetype>")
def download_member_data(filetype):
    uid = request.args.get("uid")
    if not uid:
        return jsonify({"error": "UID is required"}), 400

    member_data = get_member_data(uid)
    if not member_data:
        return jsonify({"error": "User not found"}), 404

    # Create dynamic filename
    now_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"KADHI_SACCO_MINI_Statement_as_at_{now_str}.{filetype}"

    # ---------- CSV ----------
    if filetype == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Name", "Email", "Phone", "Join Date"])
        writer.writerow([member_data["name"], member_data["email"], member_data["phone"], member_data["created_at"]])
        writer.writerow([])
        writer.writerow(["Total Contributions", sum(c["amount"] for c in member_data["contributions"])])
        writer.writerow(["Total Loans", sum(l["amount"] for l in member_data["loans"])])
        writer.writerow(["Total Repaid", sum(r["amount"] for r in member_data["repayments"])])
        writer.writerow(["Outstanding Balance", 
            sum(l["amount"] for l in member_data["loans"]) - sum(r["amount"] for r in member_data["repayments"])])
        writer.writerow(["Total Fines", sum(f["amount"] for f in member_data["fines"])])
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode("utf-8")), 
                         mimetype="text/csv", as_attachment=True, download_name=filename)

    # ---------- Excel ----------
    elif filetype == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statement"
        ws.append(["Name", "Email", "Phone", "Join Date"])
        ws.append([member_data["name"], member_data["email"], member_data["phone"], member_data["created_at"]])
        ws.append([])
        ws.append(["Total Contributions", sum(c["amount"] for c in member_data["contributions"])])
        ws.append(["Total Loans", sum(l["amount"] for l in member_data["loans"])])
        ws.append(["Total Repaid", sum(r["amount"] for r in member_data["repayments"])])
        ws.append(["Outstanding Balance", 
            sum(l["amount"] for l in member_data["loans"]) - sum(r["amount"] for r in member_data["repayments"])])
        ws.append(["Total Fines", sum(f["amount"] for f in member_data["fines"])])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)

    # ---------- PDF ----------
    elif filetype == "pdf":
        bio = io.BytesIO()
        doc = SimpleDocTemplate(bio)
        styles = getSampleStyleSheet()
        content = []
        content.append(Paragraph("<b>Member Profile Info</b>", styles["Heading2"]))
        content.append(Paragraph(f"Name: {member_data['name']}", styles["Normal"]))
        content.append(Paragraph(f"Email: {member_data['email']}", styles["Normal"]))
        content.append(Paragraph(f"Phone: {member_data['phone']}", styles["Normal"]))
        content.append(Paragraph(f"Join Date: {member_data['created_at']}", styles["Normal"]))
        content.append(Spacer(1, 12))

        content.append(Paragraph("<b>Financial Summary</b>", styles["Heading2"]))
        summary = [
            ["Total Contributions", sum(c["amount"] for c in member_data["contributions"])],
            ["Total Loans", sum(l["amount"] for l in member_data["loans"])],
            ["Total Repaid", sum(r["amount"] for r in member_data["repayments"])],
            ["Outstanding Balance", sum(l["amount"] for l in member_data["loans"]) - sum(r["amount"] for r in member_data["repayments"])],
            ["Total Fines", sum(f["amount"] for f in member_data["fines"])]
        ]
        content.append(Table(summary))
        doc.build(content)
        bio.seek(0)
        return send_file(bio, mimetype="application/pdf", as_attachment=True, download_name=filename)

    return jsonify({"error": "Invalid file type"}), 400

def get_db_connection():
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="@Dennis1043",  # replace if needed
        database="kadhi"
    )
    return conn

# --- Admin: Download All Members as CSV ---
@app.route("/download_all/csv")
def download_all_csv():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users")
    members = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    # Header
    writer.writerow([
        "Name", "Email", "Phone", "Joined", 
        "Total Contributions", "Total Loans", "Total Repaid", "Outstanding", "Total Fines"
    ])

    for member in members:
        user_id = member["id"]
        # Fetch financial summary
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT SUM(amount) as total FROM contributions WHERE user_id=%s", (user_id,))
        total_contrib = cursor.fetchone()["total"] or 0
        cursor.execute("SELECT SUM(amount) as total FROM loans WHERE user_id=%s", (user_id,))
        total_loans = cursor.fetchone()["total"] or 0
        cursor.execute("SELECT SUM(amount) as total FROM repayments WHERE user_id=%s", (user_id,))
        total_repaid = cursor.fetchone()["total"] or 0
        cursor.execute("SELECT SUM(amount) as total FROM fines WHERE user_id=%s", (user_id,))
        total_fines = cursor.fetchone()["total"] or 0
        conn.close()

        outstanding = total_loans - total_repaid

        writer.writerow([
            member["name"], member["email"], member["phone"], member["created_at"],
            total_contrib, total_loans, total_repaid, outstanding, total_fines
        ])

    output.seek(0)
    filename = f"KADHI_SACCO_Admin_Summary_as_at_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.csv"
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename
    )

# --- Admin: Download All Members as PDF ---
from flask import send_file
from fpdf import FPDF
import io
from datetime import datetime

@app.route("/download_all/pdf")
def download_all_pdf():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM users")
    members = cursor.fetchall()

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "KADHI SACCO - Admin Summary", ln=True, align="C")
    pdf.ln(5)

    now_str = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 8, f"Generated on: {now_str}", ln=True)
    pdf.ln(5)

    for member in members:
        user_id = member["id"]

        # Member info
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, f"Member: {member['name']} ({member['email']})", ln=True)
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 6, f"Phone: {member['phone']}", ln=True)
        pdf.cell(0, 6, f"Joined: {member['created_at']}", ln=True)
        pdf.ln(2)

        # Contributions
        cursor.execute("SELECT * FROM contributions WHERE user_id=%s", (user_id,))
        contributions = cursor.fetchall()
        total_contrib = sum(c["amount"] for c in contributions)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 6, f"Contributions (Total: KES {total_contrib:.2f})", ln=True)
        pdf.set_font("Arial", "", 10)
        for c in contributions:
            pdf.cell(0, 5, f"- {c['amount']:.2f} | {c['purpose']} | {c['mpesa_code']} | {c['contribution_date']}", ln=True)
        pdf.ln(1)

        # Loans
        cursor.execute("SELECT * FROM loans WHERE user_id=%s", (user_id,))
        loans = cursor.fetchall()
        total_loans = sum(l["amount"] for l in loans)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 6, f"Loans Taken (Total: KES {total_loans:.2f})", ln=True)
        pdf.set_font("Arial", "", 10)
        for l in loans:
            pdf.cell(0, 5, f"- {l['amount']:.2f} | {l['purpose']} | {l['months']} months | {l['status']} | {l['created_at']}", ln=True)
        pdf.ln(1)

        # Repayments
        cursor.execute("SELECT * FROM repayments WHERE user_id=%s", (user_id,))
        repayments = cursor.fetchall()
        total_repaid = sum(r["amount"] for r in repayments)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 6, f"Repayments (Total: KES {total_repaid:.2f})", ln=True)
        pdf.set_font("Arial", "", 10)
        for r in repayments:
            pdf.cell(0, 5, f"- {r['amount']:.2f} | {r.get('note','')} | {r.get('mpesaCode','')} | {r['timestamp']}", ln=True)
        pdf.ln(1)

        # Fines
        cursor.execute("SELECT * FROM fines WHERE user_id=%s", (user_id,))
        fines = cursor.fetchall()
        total_fines = sum(f["amount"] for f in fines)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(0, 6, f"Fines (Total: KES {total_fines:.2f})", ln=True)
        pdf.set_font("Arial", "", 10)
        for f in fines:
            pdf.cell(0, 5, f"- {f['amount']:.2f} | {f['reason']} | {f['status']} | {f['issued_on']}", ln=True)
        pdf.ln(5)

    conn.close()

    # Get PDF as bytes
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_buffer = io.BytesIO(pdf_bytes)

    filename = f"KADHI_SACCO_Admin_Summary_as_at_{now_str}.pdf"
    return send_file(pdf_buffer, download_name=filename, as_attachment=True)

import pandas as pd
from flask import send_file
import io
from datetime import datetime

@app.route("/download_all/excel")
def download_all_excel():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT * FROM users")
    members = cursor.fetchall()
    
    # Flatten members info into a list of dicts
    data = []
    for member in members:
        user_id = member['id']
        
        # Contributions
        cursor.execute("SELECT * FROM contributions WHERE user_id=%s", (user_id,))
        contributions = cursor.fetchall()
        for c in contributions:
            data.append({
                "Member Name": member["name"],
                "Email": member["email"],
                "Phone": member["phone"],
                "Type": "Contribution",
                "Amount": c["amount"],
                "Purpose": c.get("purpose", ""),
                "Code": c.get("mpesa_code", ""),
                "Date": c.get("contribution_date", "")
            })
        
        # Loans
        cursor.execute("SELECT * FROM loans WHERE user_id=%s", (user_id,))
        loans = cursor.fetchall()
        for l in loans:
            data.append({
                "Member Name": member["name"],
                "Email": member["email"],
                "Phone": member["phone"],
                "Type": "Loan",
                "Amount": l["amount"],
                "Purpose": l.get("purpose",""),
                "Months": l.get("months",""),
                "Status": l.get("status",""),
                "Date": l.get("created_at","")
            })
        
        # Fines
        cursor.execute("SELECT * FROM fines WHERE user_id=%s", (user_id,))
        fines = cursor.fetchall()
        for f in fines:
            data.append({
                "Member Name": member["name"],
                "Email": member["email"],
                "Phone": member["phone"],
                "Type": "Fine",
                "Amount": f["amount"],
                "Reason": f.get("reason",""),
                "Status": f.get("status",""),
                "Date": f.get("issued_on","")
            })

    conn.close()
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    df.to_excel(output, index=False, sheet_name="Admin Summary")
    output.seek(0)

    now_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"KADHI_SACCO_Admin_Summary_as_at_{now_str}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)


# ========================
# FRONTEND ROUTES
# ========================
@app.route("/")
def serve_index():
    return send_from_directory(".", "index.html")

@app.route("/<path:filename>")
def serve_static_files(filename):
    return send_from_directory(".", filename)

# ========================
# USER & AUTH
# ========================
@app.route("/register", methods=["POST"])
def register_user():
    data = request.json
    required_fields = ["firebase_uid", "name", "phone", "email", "role"]
    if not all(field in data for field in required_fields):
        return jsonify({"error": "Missing fields"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        sql = """
        INSERT INTO users (firebase_uid, name, phone, email, role)
        VALUES (%s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE 
            name=%s, phone=%s, email=%s, role=%s
        """
        values = (
            data["firebase_uid"], data["name"], data["phone"], data["email"], data["role"],
            data["name"], data["phone"], data["email"], data["role"]
        )
        cursor.execute(sql, values)
        conn.commit()

        cursor.close()
        conn.close()
        return jsonify({"success": True, "message": "User registered/updated"}), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/save_fcm", methods=["POST"])
def save_fcm():
    data = request.json
    if "firebase_uid" not in data or "token" not in data:
        return jsonify({"error": "Missing fields"}), 400

    try:
        user_id = get_user_id_from_firebase(data["firebase_uid"])
        if not user_id:
            return jsonify({"error": "User not found"}), 404

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT IGNORE INTO fcm_tokens (user_id, token) VALUES (%s, %s)", (user_id, data["token"]))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True, "message": "Token saved"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/user", methods=["GET"])
def get_user():
    firebase_uid = request.args.get("uid")
    if not firebase_uid:
        return jsonify({"error": "uid required"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT name, role FROM users WHERE firebase_uid=%s", (firebase_uid,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()
        if not user:
            return jsonify({"error": "User not found"}), 404
        return jsonify(user)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

# ✅ FIXED: renamed to avoid duplicate function name
@app.route("/get_user/<firebase_uid>", methods=["GET"])
def get_user_by_uid(firebase_uid):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM users WHERE firebase_uid = %s", (firebase_uid,))
    user = cursor.fetchone()

    cursor.close()
    conn.close()

    if not user:
        return jsonify({"error": "User not found"}), 404

    return jsonify(user), 200



# ========================
# CONTRIBUTIONS
# ========================
@app.route("/contributions/add", methods=["POST"])
def add_contribution():
    try:
        user_uid = request.form.get("user_id")
        amount = request.form.get("amount")
        purpose = request.form.get("purpose")
        mpesa_code = request.form.get("mpesaCode")
        screenshot = request.files.get("screenshot")  # optional

        if not user_uid or not amount:
            return jsonify({"error": "user_id and amount are required"}), 400

        user_id = get_user_id_from_firebase(user_uid)
        if not user_id:
            return jsonify({"error": "User not found"}), 404

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO contributions (user_id, amount, contribution_date, purpose, mpesa_code, status)
            VALUES (%s, %s, CURDATE(), %s, %s, 'pending')
        """, (user_id, amount, purpose, mpesa_code))
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"message": "Contribution submitted successfully"}), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/contributions/my")
def my_contributions():
    firebase_uid = request.args.get("uid")
    if not firebase_uid:
        return jsonify({"error": "Missing uid"}), 400

    user_id = get_user_id_from_firebase(firebase_uid)
    if not user_id:
        return jsonify({"error": "User not found"}), 404

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT c.id, c.amount, c.purpose, c.mpesa_code, c.status, c.timestamp, u.name AS member_name
        FROM contributions c
        JOIN users u ON c.user_id = u.id
        WHERE c.user_id=%s
        ORDER BY c.timestamp DESC
    """, (user_id,))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify({"rows": rows})

@app.route("/contributions/all")
def all_contributions():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT c.id, c.amount, c.purpose, c.mpesa_code, c.status, c.timestamp, u.name AS member_name
        FROM contributions c
        JOIN users u ON c.user_id = u.id
        ORDER BY c.timestamp DESC
    """)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify({"rows": rows})

@app.route("/contributions/<int:contrib_id>/status", methods=["POST"])
def update_contribution_status(contrib_id):
    status = request.json.get("status")
    if status not in ["approved", "rejected"]:
        return jsonify({"error": "Invalid status"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE contributions SET status=%s WHERE id=%s", (status, contrib_id))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"message": f"Contribution {status}"})

# ========================
# LOANS
# ========================
@app.route("/api/loans/request", methods=["POST"])
def request_loan():
    try:
        data = request.json
        firebase_uid = data.get("firebase_uid")
        if not firebase_uid:
            return jsonify({"error": "firebase_uid is required"}), 400

        # 🔍 Find MySQL user_id using Firebase UID
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name, phone FROM users WHERE firebase_uid = %s", (firebase_uid,))
        user = cursor.fetchone()

        if not user:
            cursor.close()
            conn.close()
            return jsonify({"error": "User not found"}), 404

        user_id = user["id"]
        name = data.get("name") or user["name"]
        phone = data.get("phone") or user["phone"]
        amount = data.get("amount")
        purpose = data.get("purpose")
        months = data.get("months")

        # Validate inputs
        if not all([amount, purpose, months]):
            return jsonify({"error": "Missing required fields"}), 400

        cursor.execute("""
            INSERT INTO loans (user_id, name, phone, amount, purpose, months, status, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, 'pending', NOW())
        """, (user_id, name, phone, amount, purpose, months))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"message": "Loan request submitted successfully"}), 201

    except Exception as e:
        print("Error in request_loan:", e)
        return jsonify({"error": str(e)}), 500


@app.route("/api/loans/repay", methods=["POST"])
def repay_loan():
    try:
        data = request.json
        firebase_uid = data.get("firebase_uid")
        if not firebase_uid:
            return jsonify({"error": "firebase_uid is required"}), 400

        # 🔍 Lookup MySQL user_id from Firebase UID
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name, phone FROM users WHERE firebase_uid = %s", (firebase_uid,))
        user = cursor.fetchone()

        if not user:
            cursor.close()
            conn.close()
            return jsonify({"error": "User not found"}), 404

        user_id = user["id"]
        phone = data.get("phone") or user["phone"]
        amount = data.get("amount")
        mpesa_code = data.get("mpesaCode")
        note = data.get("note", "")

        # Validate inputs
        if not all([amount, mpesa_code]):
            return jsonify({"error": "Missing amount or Mpesa code"}), 400

        # ✅ FIX: INSERT repayment instead of SELECT
        cursor.execute("""
            INSERT INTO repayments (user_id, phone, amount, mpesaCode, note, status, timestamp)
            VALUES (%s, %s, %s, %s, %s, 'pending', NOW())
        """, (user_id, phone, amount, mpesa_code, note))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"message": "Repayment submitted successfully"}), 201

    except Exception as e:
        print("Error in repay_loan:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/loans/history/<firebase_uid>", methods=["GET"])
def loan_history(firebase_uid):
    try:
        conn = get_mysql_connection()
        cursor = conn.cursor(dictionary=True)

        # 1️⃣ Get user_id from firebase_uid
        cursor.execute("SELECT id FROM users WHERE firebase_uid = %s", (firebase_uid,))
        user = cursor.fetchone()
        if not user:
            return jsonify({"error": "User not found"}), 404
        user_id = user["id"]

        # 2️⃣ Get loan history
        cursor.execute("""
            SELECT l.id, l.amount, l.purpose, l.status, l.created_at AS timestamp
            FROM loans l
            WHERE l.user_id = %s
            ORDER BY l.created_at DESC
        """, (user_id,))
        loans = cursor.fetchall()

        # 3️⃣ Get repayment history (✅ FIXED payment_date → timestamp)
        cursor.execute("""
            SELECT r.id, r.amount, r.mpesaCode AS mpesa_code, r.phone, r.status, r.timestamp
            FROM repayments r
            WHERE r.user_id = %s
            ORDER BY r.timestamp DESC
        """, (user_id,))
        repayments = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify({
            "loans": loans,
            "repayments": repayments
        }), 200

    except Exception as e:
        print("Error in loan_history:", e)
        return jsonify({"error": str(e)}), 500

# ========================
# FINES
# ========================
@app.route("/user-fines/<firebase_uid>", methods=["GET"])
def user_fines(firebase_uid):
    user_id = get_user_id_from_firebase(firebase_uid)
    if not user_id:
        return jsonify({"error": "User not found"}), 404

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM fines WHERE user_id=%s ORDER BY timestamp DESC", (user_id,))
    fines = cursor.fetchall()
    cursor.close()
    conn.close()

    for f in fines:
        if "timestamp" in f and f["timestamp"]:
            f["timestamp"] = f["timestamp"].strftime("%Y-%m-%d %H:%M:%S")

    return jsonify(fines)

# ========================
# SUMMARY & ADMIN
# ========================
# ========================
# ADMIN ROUTES
# ========================

# Admin: Totals (Contributions & Loans)
@app.route("/admin/totals", methods=["GET"])
def admin_totals():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT SUM(amount) as total FROM contributions WHERE status='approved'")
    contrib_total = cursor.fetchone()["total"] or 0

    cursor.execute("SELECT SUM(amount) as total FROM loans WHERE status='approved'")
    loan_total = cursor.fetchone()["total"] or 0

    cursor.close()
    conn.close()
    return jsonify({"contributions": contrib_total, "loans": loan_total})

# Admin: Members Summary

# ========================
# ADMIN FINES
# ========================
@app.route("/admin/fines/all", methods=["GET"])
def admin_fines_all():
    """Return all fines with member info."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT f.id, f.user_id, f.amount, f.reason, f.status, f.issued_on AS timestamp,
                   u.name AS member_name, u.email, u.phone
            FROM fines f
            JOIN users u ON f.user_id = u.id
            ORDER BY f.issued_on DESC
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({"fines": rows})
    except Exception as e:
        print("Error fetching fines:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/admin/members/all")
def admin_members_all():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Select users
    cursor.execute("SELECT id, name, email, phone, created_at FROM users")
    members = cursor.fetchall()

    # Fetch totals and detailed histories for each member
    for member in members:
        user_id = member['id']

        # Contributions total & history
        cursor.execute("SELECT SUM(amount) as total FROM contributions WHERE user_id=%s", (user_id,))
        member['totalContribution'] = cursor.fetchone()['total'] or 0

        cursor.execute("""
            SELECT amount, purpose, mpesa_code, status, contribution_date, timestamp
            FROM contributions
            WHERE user_id=%s
            ORDER BY timestamp DESC
        """, (user_id,))
        member['contributions'] = cursor.fetchall()

        # Loans total & history
        cursor.execute("SELECT SUM(amount) as total FROM loans WHERE user_id=%s", (user_id,))
        member['totalLoan'] = cursor.fetchone()['total'] or 0

        cursor.execute("""
            SELECT amount, purpose, months, status, created_at
            FROM loans
            WHERE user_id=%s
            ORDER BY created_at DESC
        """, (user_id,))
        member['loans'] = cursor.fetchall()

        # Fines total & history
        cursor.execute("SELECT SUM(amount) as total FROM fines WHERE user_id=%s", (user_id,))
        member['totalFines'] = cursor.fetchone()['total'] or 0

        cursor.execute("""
            SELECT amount, reason, status, issued_on
            FROM fines
            WHERE user_id=%s
            ORDER BY issued_on DESC
        """, (user_id,))
        member['fines'] = cursor.fetchall()

    cursor.close()
    conn.close()
    return jsonify({"members": members})


@app.route("/balance", methods=["GET"])
def user_balance():
    firebase_uid = request.args.get("uid")
    if not firebase_uid:
        return jsonify({"error": "uid required"}), 400

    user_id = get_user_id_from_firebase(firebase_uid) 
    if not user_id:
        return jsonify({"error": "User not found"}), 404

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Total contributions
    cursor.execute("SELECT SUM(amount) AS total FROM contributions WHERE user_id=%s AND status='approved'", (user_id,))
    contributions = cursor.fetchone()["total"] or 0

    # Total loans taken
    cursor.execute("SELECT SUM(amount) AS total FROM loans WHERE user_id=%s AND status='approved'", (user_id,))
    loans = cursor.fetchone()["total"] or 0

    # Total repayments
    cursor.execute("SELECT SUM(amount) AS total FROM repayments WHERE user_id=%s AND status='approved'", (user_id,))
    repayments = cursor.fetchone()["total"] or 0

    cursor.close()
    conn.close()

    return jsonify({
        "contributions": contributions,
        "loans": loans,
        "repayments": repayments
    })


# Admin: Notifications
# ========================
# ADMIN NOTIFICATIONS
# ========================
@app.route("/admin/notifications", methods=["GET"])
def admin_notifications():
    """Return all pending notifications for admin dashboard."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Pending contributions
        cursor.execute("""
            SELECT c.id, c.amount, c.user_id, u.name AS member_name
            FROM contributions c
            JOIN users u ON c.user_id = u.id
            WHERE c.status='pending'
            ORDER BY c.timestamp DESC
        """)
        contributions = cursor.fetchall()

        # Pending loan requests
        cursor.execute("""
            SELECT l.id, l.amount, l.user_id, u.name AS member_name
            FROM loans l
            JOIN users u ON l.user_id = u.id
            WHERE l.status='pending'
            ORDER BY l.created_at DESC
        """)
        loans = cursor.fetchall()

        # Pending repayments
        cursor.execute("""
            SELECT r.id, r.amount, r.user_id, u.name AS member_name
            FROM repayments r
            JOIN users u ON r.user_id = u.id
            WHERE r.status='pending'
            ORDER BY r.timestamp DESC
        """)
        repayments = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify({
            "contributions": contributions,
            "loans": loans,
            "repayments": repayments
        })

    except Exception as e:
        print("Error fetching admin notifications:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/admin/loans/all", methods=["GET"])
def admin_all_loans():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT l.id, l.amount, l.purpose, l.status, l.created_at AS timestamp, u.name AS member_name
        FROM loans l
        JOIN users u ON l.user_id = u.id
        ORDER BY l.created_at DESC
    """)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify({"rows": rows})

@app.route("/admin/loans/<int:loan_id>/status", methods=["POST"])
def admin_update_loan_status(loan_id):
    status = request.json.get("status")
    if status not in ["approved", "rejected"]:
        return jsonify({"error": "Invalid status"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE loans SET status=%s WHERE id=%s", (status, loan_id))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"message": f"Loan {status}"})


# ========================
# ADMIN REPAYMENTS
# ========================
@app.route("/admin/repayments/all", methods=["GET"])
def admin_repayments_all():
    """Return all repayments (pending and approved) with member names."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT r.id, r.user_id, r.amount, r.note, r.phone, r.mpesaCode AS mpesa_code,
                   r.status, r.timestamp, u.name AS member_name
            FROM repayments r
            JOIN users u ON r.user_id = u.id
            ORDER BY r.timestamp DESC
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({"rows": rows})
    except Exception as e:
        print("Error fetching admin repayments:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/admin/repayments/<int:repay_id>/status", methods=["POST"])
def admin_update_repayment_status(repay_id):
    """Approve or reject a repayment."""
    try:
        status = request.json.get("status")
        if status not in ["approved", "rejected"]:
            return jsonify({"error": "Invalid status"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE repayments SET status=%s WHERE id=%s", (status, repay_id))
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"message": f"Repayment {status}"})
    except Exception as e:
        print("Error updating repayment status:", e)
        return jsonify({"error": str(e)}), 500

# ========================
# SETTINGS
# ========================
@app.route("/settings", methods=["GET"])
def get_settings():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM settings WHERE id=1")
    row = cursor.fetchone()
    cursor.close(); conn.close()
    return jsonify(row if row else {})

@app.route("/settings", methods=["POST"])
def save_settings():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO settings 
        (id, monthly_contribution, loan_interest, default_fine, grace_period, contribution_duration, next_contribution_date)
        VALUES (1, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
        monthly_contribution=VALUES(monthly_contribution),
        loan_interest=VALUES(loan_interest),
        default_fine=VALUES(default_fine),
        grace_period=VALUES(grace_period),
        contribution_duration=VALUES(contribution_duration),
        next_contribution_date=VALUES(next_contribution_date)
    """, (
        data.get("monthlyContribution"),
        data.get("loanInterest"),
        data.get("defaultFine"),
        data.get("gracePeriod"),
        data.get("contributionDuration"),
        data.get("nextContributionDate")
    ))
    conn.commit()
    cursor.close(); conn.close()
    return jsonify({"message": "Settings saved successfully"})

@app.route("/notifications", methods=["GET"])
def get_user_notifications():
    firebase_uid = request.args.get("uid")
    if not firebase_uid:
        return jsonify({"error": "uid required"}), 400
    try:
        user_id = get_user_id_from_firebase(firebase_uid)
        if not user_id:
            return jsonify({"error": "User not found"}), 404
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, message, read_flag
            FROM notifications
            WHERE user_id=%s AND read_flag=0
        """, (user_id,))
        notifications = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({"rows": notifications})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ========================
# ADMIN: IMPORT MEMBERS DATA
# ========================
import pandas as pd
import firebase_admin
from firebase_admin import auth, credentials
import random, string

# Initialize Firebase Admin (update path to your service account)
try:
    firebase_admin.get_app()
except ValueError:
    cred = credentials.Certificate("kadhi-service-account.json")
    firebase_admin.initialize_app(cred)

@app.route("/admin/import_data", methods=["POST"])
def admin_import_data():
    if "file" not in request.files:
        return jsonify({"error":"No file uploaded"}), 400

    file = request.files["file"]
    try:
        if file.filename.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
    except Exception as e:
        return jsonify({"error": f"Failed to read file: {str(e)}"}), 400

    required_columns = ["name","phone","email","contribution","loan_amount","repayment","fine"]
    for col in required_columns:
        if col not in df.columns:
            return jsonify({"error": f"Missing column: {col}"}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    errors = []
    imported_count = 0

    for idx, row in df.iterrows():
        try:
            name = str(row["name"]).strip()
            phone = str(row["phone"]).strip()
            email = str(row["email"]).strip()
            contribution = float(row["contribution"] or 0)
            loan_amount = float(row["loan_amount"] or 0)
            repayment = float(row["repayment"] or 0)
            fine = float(row["fine"] or 0)

            # Check if user exists in MySQL
            cursor.execute("SELECT id, firebase_uid FROM users WHERE email=%s", (email,))
            user = cursor.fetchone()
            if user:
                user_id, firebase_uid = user["id"], user["firebase_uid"]
            else:
                # Check if user exists in Firebase
                try:
                    fb_user = auth.get_user_by_email(email)
                    firebase_uid = fb_user.uid
                except auth.UserNotFoundError:
                    # Create new Firebase user
                    default_password = "Password123!"
                    random_uid = ''.join(random.choices(string.ascii_letters+string.digits, k=28))
                    fb_user = auth.create_user(uid=random_uid, email=email, password=default_password, display_name=name)
                    firebase_uid = fb_user.uid

                # Insert user into MySQL
                cursor.execute("""
                    INSERT INTO users (firebase_uid, name, email, phone, role)
                    VALUES (%s,%s,%s,%s,'member')
                """, (firebase_uid, name, email, phone))
                conn.commit()
                user_id = cursor.lastrowid

            # Insert contribution
            if contribution > 0:
                cursor.execute("""
                    INSERT INTO contributions (user_id, amount, contribution_date, status)
                    VALUES (%s,%s,CURDATE(),'approved')
                """, (user_id, contribution))

            # Insert loan
            if loan_amount > 0:
                cursor.execute("""
                    INSERT INTO loans (user_id, name, phone, amount, purpose, months, status, created_at)
                    VALUES (%s,%s,%s,%s,'Imported',12,'approved',NOW())
                """, (user_id, name, phone, loan_amount))

            # Insert repayment
            if repayment > 0:
                cursor.execute("""
                    INSERT INTO repayments (user_id, phone, amount, mpesaCode, note, status, timestamp)
                    VALUES (%s,%s,%s,'Imported','Imported','approved',NOW())
                """, (user_id, phone, repayment))

            # Insert fine
            if fine > 0:
                cursor.execute("""
                    INSERT INTO fines (user_id, amount, reason, status, issued_on)
                    VALUES (%s,%s,'Imported','paid',NOW())
                """, (user_id, fine))

            conn.commit()
            imported_count += 1

        except Exception as e:
            conn.rollback()
            errors.append(f"Row {idx+2}: {str(e)}")
            continue  # continue with next row

    cursor.close()
    conn.close()

    result = {"message": f"Import finished. {imported_count} rows imported successfully."}
    if errors:
        result["errors"] = errors

    return jsonify(result)


from werkzeug.security import generate_password_hash, check_password_hash

# ========================
# PASSWORD MANAGEMENT
# ========================

@app.route("/user/password/change", methods=["POST"])
def change_password():
    """
    Allows a user to change their password.
    Expects JSON:
    {
        "firebase_uid": "...",
        "current_password": "...",  # optional if implementing reset via token
        "new_password": "..."
    }
    """
    data = request.json
    firebase_uid = data.get("firebase_uid")
    new_password = data.get("new_password")
    current_password = data.get("current_password")

    if not firebase_uid or not new_password:
        return jsonify({"error": "firebase_uid and new_password are required"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Fetch current password hash from DB
        cursor.execute("SELECT id, password_hash FROM users WHERE firebase_uid=%s", (firebase_uid,))
        user = cursor.fetchone()
        if not user:
            cursor.close()
            conn.close()
            return jsonify({"error": "User not found"}), 404

        user_id = user["id"]
        stored_hash = user.get("password_hash")

        # Optional: verify current password if provided
        if current_password and stored_hash:
            if not check_password_hash(stored_hash, current_password):
                cursor.close()
                conn.close()
                return jsonify({"error": "Current password is incorrect"}), 401

        # Hash new password and update
        new_hash = generate_password_hash(new_password)
        cursor.execute("UPDATE users SET password_hash=%s WHERE id=%s", (new_hash, user_id))
        conn.commit()

        cursor.close()
        conn.close()
        return jsonify({"message": "Password updated successfully"}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


from werkzeug.security import generate_password_hash

@app.route("/user/password/reset", methods=["POST"])
def reset_password():
    """Reset user password in MySQL and Firebase"""
    data = request.json
    email = data.get("email")
    firebase_uid = data.get("firebase_uid")
    new_password = data.get("new_password")

    if not new_password or (not email and not firebase_uid):
        return jsonify({"error": "Email or firebase_uid and new_password are required"}), 400

    try:
        # 1️⃣ Lookup user in MySQL
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        if email:
            cursor.execute("SELECT id, firebase_uid FROM users WHERE email=%s", (email,))
        else:
            cursor.execute("SELECT id, firebase_uid FROM users WHERE firebase_uid=%s", (firebase_uid,))

        user = cursor.fetchone()
        if not user:
            cursor.close()
            conn.close()
            return jsonify({"error": "User not found"}), 404

        user_id = user["id"]
        firebase_uid = user["firebase_uid"]

        # 2️⃣ Update password in MySQL
        cursor.execute("UPDATE users SET password=%s WHERE id=%s", (new_password, user_id))
        conn.commit()
        cursor.close()
        conn.close()

        # 3️⃣ Update password in Firebase
        auth.update_user(firebase_uid, password=new_password)

        return jsonify({"message": "Password reset successfully"}), 200

    except Exception as e:
        print("Error resetting password:", e)
        return jsonify({"error": str(e)}), 500
    
@app.route("/admin/reset_all_passwords", methods=["POST"])
def reset_all_passwords():
    """Reset all users' passwords to a default value for first login."""
    DEFAULT_PASSWORD = "Password123!"  # Everyone gets this temporarily

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Fetch all users
        cursor.execute("SELECT id, firebase_uid, email FROM users")
        users = cursor.fetchall()

        if not users:
            cursor.close()
            conn.close()
            return jsonify({"message": "No users found"}), 404

        updated_count = 0
        errors = []

        for user in users:
            user_id = user["id"]
            firebase_uid = user["firebase_uid"]

            try:
                # Update MySQL
                cursor.execute("UPDATE users SET password=%s WHERE id=%s", (DEFAULT_PASSWORD, user_id))
                conn.commit()

                # Update Firebase
                auth.update_user(firebase_uid, password=DEFAULT_PASSWORD)
                updated_count += 1

            except Exception as e:
                errors.append(f"{user['email']}: {str(e)}")
                continue

        cursor.close()
        conn.close()
        return jsonify({
            "message": f"{updated_count} users reset to default password.",
            "errors": errors
        }), 200

    except Exception as e:
        print("Error resetting all passwords:", e)
        return jsonify({"error": str(e)}), 500


# ========================
# GET USER BY EMAIL
# ========================
@app.route("/get_user_by_email", methods=["GET"])
def get_user_by_email():
    email = request.args.get("email")
    if not email:
        return jsonify({"error": "Email is required"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name, phone, firebase_uid FROM users WHERE email=%s", (email,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if not user:
            return jsonify({"error": "User not found"}), 404

        return jsonify(user), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
    
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

@app.route("/download/my_data_pdf", methods=["GET"])
def download_my_data_pdf():
    """
    Download a professional PDF statement for a member.
    Query params:
    - uid: Firebase UID of the user
    """
    firebase_uid = request.args.get("uid")
    if not firebase_uid:
        return jsonify({"error": "uid required"}), 400

    user_id = get_user_id_from_firebase(firebase_uid)
    if not user_id:
        return jsonify({"error": "User not found"}), 404

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Member info
    cursor.execute("SELECT name, email, phone FROM users WHERE id=%s", (user_id,))
    member = cursor.fetchone()

    # Fetch all relevant data
    cursor.execute("""SELECT amount, purpose, mpesa_code, status, contribution_date
                      FROM contributions WHERE user_id=%s ORDER BY contribution_date DESC""", (user_id,))
    contributions = cursor.fetchall()

    cursor.execute("""SELECT amount, purpose, months, status, created_at
                      FROM loans WHERE user_id=%s ORDER BY created_at DESC""", (user_id,))
    loans = cursor.fetchall()

    cursor.execute("""SELECT amount, mpesaCode AS mpesa_code, phone, status, timestamp
                      FROM repayments WHERE user_id=%s ORDER BY timestamp DESC""", (user_id,))
    repayments = cursor.fetchall()

    cursor.execute("""SELECT amount, reason, status, issued_on
                      FROM fines WHERE user_id=%s ORDER BY issued_on DESC""", (user_id,))
    fines = cursor.fetchall()

    cursor.close()
    conn.close()

    # Calculate totals
    total_contrib = sum(c['amount'] for c in contributions) if contributions else 0
    total_loans = sum(l['amount'] for l in loans) if loans else 0
    total_repayments = sum(r['amount'] for r in repayments) if repayments else 0
    total_fines = sum(f['amount'] for f in fines) if fines else 0

    # Prepare PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(f"Member Statement", styles['Title']))
    elements.append(Spacer(1, 12))

    # Member info
    elements.append(Paragraph(f"<b>Name:</b> {member['name']}", styles['Normal']))
    elements.append(Paragraph(f"<b>Email:</b> {member['email']}", styles['Normal']))
    elements.append(Paragraph(f"<b>Phone:</b> {member['phone']}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Totals
    elements.append(Paragraph(f"<b>Total Contributions:</b> {total_contrib}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Loans:</b> {total_loans}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Repayments:</b> {total_repayments}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Fines:</b> {total_fines}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Helper to create tables
    def make_table(title, data, columns):
        if not data:
            elements.append(Paragraph(f"<b>{title}:</b> No records", styles['Normal']))
            elements.append(Spacer(1, 12))
            return
        elements.append(Paragraph(f"<b>{title}</b>", styles['Heading3']))
        table_data = [columns] + [[row[col] for col in columns] for row in data]
        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige)
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))

    # Add tables
    make_table("Contributions", contributions, ["amount", "purpose", "mpesa_code", "status", "contribution_date"])
    make_table("Loans", loans, ["amount", "purpose", "months", "status", "created_at"])
    make_table("Repayments", repayments, ["amount", "mpesa_code", "phone", "status", "timestamp"])
    make_table("Fines", fines, ["amount", "reason", "status", "issued_on"])

    # Build PDF
    doc.build(elements)
    output.seek(0)

    return send_file(output, mimetype="application/pdf",
                     download_name=f"{member['name']}_statement.pdf", as_attachment=True)


# ========================
# START SERVER
# ========================
if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)


